import openpyxl
import sys

fileName = sys.argv[1]
wb = openpyxl.load_workbook(fileName)
sheet = wb.active
all_data = []
for row in range(1, sheet.max_row+1):
    entire_row = []
    for column in range(1, sheet.max_column+1):
        entire_row.append(sheet.cell(row=row, column=column).value)
        sheet.cell(row=row, column=column).value = ''
    all_data.append(entire_row)
for row in range(len(all_data)):
    for column in range(len(all_data[row])):
        sheet.cell(row=column+1, column=row+1).value = all_data[row][column]

wb.save(fileName[:5] + '_inverted.xlsx')
