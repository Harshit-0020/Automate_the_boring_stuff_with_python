#! python3
# blankRowInsterter.py - inserts given number of blank rows at the specified location.

import openpyxl
from openpyxl.utils import get_column_letter
import sys

N = int(sys.argv[1])
M = int(sys.argv[2])
fileName = sys.argv[3]


# Method 1:- (Using sheet add rows method)
""" wb = openpyxl.load_workbook(fileName)
sheet = wb.active
print('Inserting the rows...')
sheet.insert_rows(idx=N, amount=M)
print('Saving the workbook...')
newFileName = fileName[:(len(fileName)-5)]+'_copy.xlsx'
wb.save(newFileName)
print(f'File saved as {newFileName}')
print('Done!') """

# Method 2:- (Using a little bit manual method)
wb = openpyxl.load_workbook(fileName)
sheet = wb.active

for rowOfCellObjects in tuple(sheet[f'A{N}':f'{get_column_letter(sheet.max_column)}{sheet.max_row}'])[::-1]:
    for cell_obj in rowOfCellObjects:
        sheet.cell(row=cell_obj.row+M, column=cell_obj.column).value = cell_obj.value
        cell_obj.value = ''

newFileName = fileName[:(len(fileName)-5)]+'_copy.xlsx'
wb.save(newFileName)
print(f'File saved as {newFileName}')
